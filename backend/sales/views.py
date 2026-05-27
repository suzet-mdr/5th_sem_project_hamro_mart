from rest_framework import viewsets, status, generics
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, F, Count
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from accounts.permissions import IsAdmin, IsManagerOrAbove, IsStaff
from .models import Sale, SaleItem, Expense, DailySummary
from .serializers import SaleSerializer, ExpenseSerializer, DailySummarySerializer

class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all().order_by('-created_at')
    serializer_class = SaleSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['invoice_number', 'customer__phone', 'customer__name']
    filterset_fields = ['payment_method', 'payment_status']

    def get_permissions(self):
        if self.request.method in ['GET', 'HEAD', 'OPTIONS']:
            return [IsStaff()]
        return [IsStaff()]  # Cashiers/Staff can record sales

    def perform_create(self, serializer):
        serializer.save(cashier=self.request.user)

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def profit_report(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        sales = Sale.objects.filter(created_at__date__range=[start_date, end_date])
        total_revenue = sales.aggregate(sum_total=Sum('total'))['sum_total'] or 0.0
        
        # total cost
        total_cost = SaleItem.objects.filter(sale__in=sales).aggregate(
            sum_cost=Sum(F('quantity') * F('cost_price'))
        )['sum_cost'] or 0.0

        expenses = Expense.objects.filter(date__range=[start_date, end_date])
        total_expenses = expenses.aggregate(sum_amount=Sum('amount'))['sum_amount'] or 0.0

        gross_profit = float(total_revenue) - float(total_cost)
        net_profit = gross_profit - float(total_expenses)

        return Response({
            "start_date": start_date,
            "end_date": end_date,
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "gross_profit": gross_profit,
            "total_expenses": total_expenses,
            "net_profit": net_profit
        })

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def export_excel(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

        sales = Sale.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Headers
        headers = ["Date", "Invoice Number", "Customer Name", "Customer Phone", "Cashier", "Payment Method", "Subtotal", "Discount", "Tax", "Total"]
        ws.append(headers)

        # Style headers
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for sale in sales:
            ws.append([
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
                sale.invoice_number,
                sale.customer.name if sale.customer else "Guest",
                sale.customer.phone if sale.customer else "-",
                sale.cashier.username if sale.cashier else "System",
                sale.get_payment_method_display(),
                sale.subtotal,
                sale.discount,
                sale.tax,
                sale.total
            ])

        # Formatting columns width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(response)
        return response

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def export_pdf(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not start_date_str or not end_date_str:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

        sales = Sale.objects.filter(created_at__date__range=[start_date, end_date]).order_by('-created_at')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Custom Styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F497D'),
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20
        )

        elements.append(Paragraph("Hamro Mart - Sales Report", title_style))
        elements.append(Paragraph(f"Reporting Period: {start_date} to {end_date}", subtitle_style))
        elements.append(Spacer(1, 10))

        # Table data
        data = [["Date", "Invoice", "Customer", "Payment", "Total"]]
        total_sales_val = 0.0
        for sale in sales:
            data.append([
                sale.created_at.strftime("%Y-%m-%d"),
                sale.invoice_number,
                sale.customer.name if sale.customer else "Guest",
                sale.get_payment_method_display(),
                f"Rs. {sale.total}"
            ])
            total_sales_val += float(sale.total)

        data.append(["", "", "", "Total Revenue:", f"Rs. {total_sales_val}"])

        table = Table(data, colWidths=[80, 100, 120, 100, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-2), colors.HexColor('#F2F2F2')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (-2,-1), (-1,-1), 'Helvetica-Bold'),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1F497D')),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all().order_by('-date')
    serializer_class = ExpenseSerializer

    def get_permissions(self):
        if self.request.method in ['GET', 'HEAD', 'OPTIONS']:
            return [IsStaff()]
        return [IsManagerOrAbove()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class ReportSummaryView(APIView):
    permission_classes = [IsStaff]

    def get(self, request):
        # returns simple trend charts endpoints
        today = datetime.now().date()
        
        # Last 7 days trend
        seven_days_ago = today - timedelta(days=7)
        daily_sales = Sale.objects.filter(created_at__date__range=[seven_days_ago, today])\
            .annotate(date=TruncDate('created_at'))\
            .values('date')\
            .annotate(revenue=Sum('total'), count=Count('id'))\
            .order_by('date')

        # Last 6 months trend
        six_months_ago = today - timedelta(days=180)
        monthly_sales = Sale.objects.filter(created_at__date__range=[six_months_ago, today])\
            .annotate(month=TruncMonth('created_at'))\
            .values('month')\
            .annotate(revenue=Sum('total'), count=Count('id'))\
            .order_by('month')

        return Response({
            "daily_trend": daily_sales,
            "monthly_trend": monthly_sales
        })
